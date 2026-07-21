r"""目录分层、双 OCR 对比和人工复核处理脚本。

* 使用前需要先修改本脚本开头的 TABLE_PATH、PDF_DIR、OUTPUT_DIR。
  表 A、表 B 都从 TABLE_PATH 读取，OCR 结果只写入 OUTPUT_DIR。
* md 文件中表格图片需要使用md阅读器进行渲染阅读，推荐使用obesdian
使用顺序：
1. outline：识别目录页、调用大模型划分目录层级，并生成
   风水书籍转换/<书名>/final_result/outline.md。完成后暂停，人工检查目录。
2. compare：先从已检查的目录中删除反选章节，再对入选正文页执行 PaddleOCR 和
   MinerU；正常页面写入 final_result，异常页面写入 review.xlsx。完成后暂停复核。
3. finalize：根据 review.xlsx 的处理方式生成最终页面、修正标题层级并输出 result.json。
   处理方式为空的行会跳过，整张表处理完后在控制台打印对应 PDF 页码。

常用命令（在本脚本所在目录运行）：
    python .\dual_ocr_pipeline.py outline
    python .\dual_ocr_pipeline.py compare
    python .\dual_ocr_pipeline.py compare --overwrite
    python .\dual_ocr_pipeline.py finalize

只处理指定书籍：
    python .\dual_ocr_pipeline.py outline --book "书名一" "书名二"
    python .\dual_ocr_pipeline.py compare --book "书名一"
    python .\dual_ocr_pipeline.py finalize --book "书名一"

    
本次不需要管的参数（已被写死）
反向筛选参数（被选中的章节不处理，四类条件之间为并集）：
    --exclude-structure "文件结构分类"
    --exclude-scene "建筑应用场景"
    --exclude-branch "建筑应用场景子分支"
    --exclude-pair "建筑应用场景-建筑应用场景子分支"
例如：
    python .\dual_ocr_pipeline.py compare --exclude-scene "商业建筑" --exclude-pair "住宅-周边环境"

其他参数：
    --dry-run / --no-dry-run   只校验和打印计划，不调用 OCR 或模型接口。
    --outline-model MODEL      指定目录分层模型，仅 outline 阶段使用。
    --overwrite                仅 compare 阶段使用：在 OCR 前清空原有 page_*.md 和
                               review.xlsx，保留 outline.md；已有完整 OCR 页面仍会跳过。

未传入反选参数或 dry-run 参数时，使用本脚本顶部的全局变量；
命令行参数的优先级高于全局变量。表 B 中对应 PDF 不存在的书籍会直接跳过。
MinerU 按每批五页并发处理，批次之间串行。
"""

import argparse
import asyncio
import base64
import csv
import html
import io
import json
import pathlib
import re
import shutil
import unicodedata
import zipfile

import httpx
import pypdfium2 as pdfium
from openai import AsyncOpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

import layout_pages_transfer as paddle


# dual 是路径的唯一来源；layout 的默认路径只供其被单独执行时使用。
TABLE_PATH = pathlib.Path("风水书籍转换")
PDF_DIR = pathlib.Path(r"D:\pythonprojects\风水图片rag测试\测试书籍")
OUTPUT_DIR = pathlib.Path("风水书籍转换")


TABLE_A_PATTERN = "1-DD-RAG应用 - A-整体进度汇总*.csv"
TABLE_B_PATH = TABLE_PATH / "1-DD-RAG应用 - B-步骤2-PDF章节分类.csv"

PADDLE_BASE_URL = "http://172.168.48.51:4002"
MINERU_URL = "http://172.168.47.52:4003/file_parse"
OUTLINE_BASE_URL = "http://172.168.48.51:4000/v1"
OUTLINE_MODEL = "qwen3.5-122b"
MINERU_BATCH_SIZE = 5
EXCLUDE_STRUCTURES = []
EXCLUDE_SCENES = ["商业建筑", "其他建筑"]
EXCLUDE_BRANCHES = ["周边环境", "公共空间"]
EXCLUDE_PAIRS = []  # 例如 [("住宅", "通用"), ("商业建筑", "周边环境")]
DRY_RUN = False
REVIEW_OPTIONS = ("保留PaddleOCR", "保留MinerU", "手动merge", "删除该页", "两种都可")

# layout 被本脚本导入后，所有共享配置统一由 dual 单向覆盖。
paddle.TABLE_PATH = TABLE_PATH
paddle.TABLE_B_PATH = TABLE_B_PATH
paddle.PDF_DIR = PDF_DIR
paddle.OUTPUT_DIR = OUTPUT_DIR
paddle.BASE_URL = PADDLE_BASE_URL
paddle.EXCLUDE_STRUCTURES = EXCLUDE_STRUCTURES
paddle.EXCLUDE_SCENES = EXCLUDE_SCENES
paddle.EXCLUDE_BRANCHES = EXCLUDE_BRANCHES
paddle.EXCLUDE_PAIRS = EXCLUDE_PAIRS
paddle.DRY_RUN = DRY_RUN

OUTLINE_RE = re.compile(
    r"^(?P<heading>#{2,6}\s+)?(?P<title>.+?)"
    r"(?:\s*[.…·•．。_—-]{2,}\s*|\s{2,})(?P<page>\d+)\s*$"
)
HEADING_RE = re.compile(r"^(?P<indent>\s*)#{1,6}\s+(?P<title>.+?)\s*$")
MEDIA_TYPES = {"image", "table", "chart", "image_body", "table_body"}
COMPARE_PUNCTUATION = str.maketrans(
    {
        "“": '"',
        "”": '"',
        "＂": '"',
        "‘": "'",
        "’": "'",
        "＇": "'",
        "（": "(",
        "）": ")",
        "［": "[",
        "］": "]",
        "【": "[",
        "】": "]",
        "｛": "{",
        "｝": "}",
        "＜": "<",
        "＞": ">",
        "，": ",",
    }
)


def open_csv_with_fallback(path):
    """优先按 UTF-8-SIG 读取 CSV，失败时回退到 GB18030。"""

    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="gb18030")
        print(f"CSV 使用 GB18030 解码：{path}")
    return io.StringIO(text, newline="")


def load_b_tasks(selected_books=None):
    """从表 B 读取正文转换任务，并校验 PDF、页码范围及逐页 metadata。"""

    with open_csv_with_fallback(TABLE_B_PATH) as f:
        rows = [
            (row_number, row)
            for row_number, row in enumerate(csv.DictReader(f), start=2)
            if row["开头章节标题"].strip() and row["属性"].strip() != "说明"
        ]

    available_books = {row["文件名称"].strip() for _, row in rows}
    if selected_books:
        unknown = set(selected_books) - available_books
        if unknown:
            raise ValueError(f"表 B 中不存在书名：{', '.join(sorted(unknown))}")
        rows = [(number, row) for number, row in rows if row["文件名称"].strip() in selected_books]

    missing_books = sorted(
        {
            row["文件名称"].strip()
            for _, row in rows
            if not (PDF_DIR / f"{row['文件名称'].strip()}.pdf").is_file()
        }
    )
    for book in missing_books:
        print(f"跳过：PDF_DIR 中不存在 {book}.pdf")
    rows = [(number, row) for number, row in rows if row["文件名称"].strip() not in missing_books]

    tasks = []
    occupied_pages = {}
    page_counts = {}
    for row_number, row in rows:
        filename = row["文件名称"].strip()
        pdf_path = PDF_DIR / f"{filename}.pdf"
        if pdf_path not in page_counts:
            with pdfium.PdfDocument(str(pdf_path)) as pdf:
                page_counts[pdf_path] = len(pdf)
        page_count = page_counts[pdf_path]

        try:
            start_page = int(row["开头页码"].strip())
            end_value = row["结尾页码"].strip()
            end_page = page_count + 1 if end_value == "-" else int(end_value)
        except ValueError as exc:
            raise ValueError(f"表 B 第 {row_number} 行页码必须是数字或 -") from exc
        if not 1 <= start_page < end_page <= page_count + 1:
            raise ValueError(
                f"表 B 第 {row_number} 行范围 [{start_page}, {end_page}) "
                f"超出 {pdf_path.name} 的 1-{page_count} 页"
            )

        pages = set(range(start_page, end_page))
        overlap = occupied_pages.setdefault(pdf_path, set()) & pages
        if overlap:
            raise ValueError(f"表 B 第 {row_number} 行页码重叠：{pdf_path.name} 第 {min(overlap)} 页")
        occupied_pages[pdf_path].update(pages)

        metadata = {
            key: row[key].strip()
            for key in ("文件结构分类", "建筑应用场景", "建筑应用场景子分支")
        }
        missing_metadata = [key for key, value in metadata.items() if not value]
        if missing_metadata:
            raise ValueError(f"表 B 第 {row_number} 行缺少 metadata：{', '.join(missing_metadata)}")
        tasks.append(
            {
                "row_number": row_number,
                "chapter": row["开头章节标题"].strip(),
                "pdf_path": pdf_path,
                "start_page": start_page,
                "end_page": end_page,
                "metadata": metadata,
            }
        )
    return tasks, missing_books


def load_book_infos(book_names):
    """从表 A 为待处理书籍读取目录页范围和目录页码偏移。"""

    candidates = list(TABLE_PATH.glob(TABLE_A_PATTERN))
    if len(candidates) != 1:
        raise FileNotFoundError(
            f"表 A 应唯一匹配 {TABLE_A_PATTERN}，实际找到："
            f"{', '.join(path.name for path in candidates) or '无'}"
        )

    with open_csv_with_fallback(candidates[0]) as f:
        rows = list(csv.reader(f))
    if len(rows) < 3:
        raise ValueError("表 A 至少需要三行表头")
    headers = rows[2]
    required = ("重命名", "目录所在页码开头", "目录所在页码结尾", "目录页码偏离")
    indices = {}
    for name in required:
        if headers.count(name) != 1:
            raise ValueError(f"表 A 字段必须唯一：{name}")
        indices[name] = headers.index(name)

    matches = {book: [] for book in book_names}
    for row in rows[3:]:
        if len(row) <= indices["重命名"]:
            continue
        name = row[indices["重命名"]].strip()
        if name in matches:
            matches[name].append(row)

    infos = {}
    for book, found in matches.items():
        if len(found) != 1:
            raise ValueError(f"表 A 的重命名={book} 应匹配 1 行，实际匹配 {len(found)} 行")
        row = found[0]
        try:
            toc_start = int(row[indices["目录所在页码开头"]].strip())
            toc_end = int(row[indices["目录所在页码结尾"]].strip())
            page_offset = int(row[indices["目录页码偏离"]].strip())
        except (ValueError, IndexError) as exc:
            raise ValueError(f"表 A 中 {book} 的目录页码或偏移不是整数") from exc
        if not 1 <= toc_start < toc_end:
            raise ValueError(f"表 A 中 {book} 的目录范围无效：[{toc_start}, {toc_end})")
        infos[book] = {
            "toc_start": toc_start,
            "toc_end": toc_end,
            "page_offset": page_offset,
        }
    return infos


def group_tasks(tasks):
    """按书名聚合同一本 PDF 的章节转换任务。"""

    grouped = {}
    for task in tasks:
        grouped.setdefault(task["pdf_path"].stem, []).append(task)
    return grouped


def page_metadata(tasks):
    """展开章节任务，生成正文 PDF 页码到 metadata 的映射。"""

    return {
        page: task["metadata"]
        for task in tasks
        for page in range(task["start_page"], task["end_page"])
    }


def normalize_title(value):
    """统一标题字符、Markdown 标记和空白，供跨 OCR 标题匹配使用。"""

    value = unicodedata.normalize("NFKC", value)
    value = re.sub(r"[`*_]", "", value)
    return re.sub(r"\s+", "", value).strip("：:。.．")


def strip_front_matter(text):
    """移除页面 Markdown 开头的 YAML metadata，保留正文内容。"""

    lines = text.splitlines()
    if lines and lines[0].strip() == "---":
        for index in range(1, len(lines)):
            if lines[index].strip() == "---":
                return "\n".join(lines[index + 1 :])
    return text


def normalized_markdown(text):
    """规范化换行、标题层级、横向空白及全半角标点，供两种 OCR 比较。"""

    lines = (
        re.sub(
            r"[\t \u3000]+",
            "",
            re.sub(
                r"^(\s*)#{1,6}(?=\s)",
                r"\1#",
                line.rstrip().translate(COMPARE_PUNCTUATION),
            ),
        )
        for line in strip_front_matter(text).replace("\r\n", "\n").split("\n")
    )
    return "\n".join(lines).rstrip("\n")


def add_metadata(text, pdf_name, metadata):
    """为单页 OCR Markdown 写入文件名和三类业务 metadata。"""

    values = {"文件名": pdf_name, **metadata}
    front_matter = "\n".join(
        f"{key}: {json.dumps(value, ensure_ascii=False)}" for key, value in values.items()
    )
    return f"---\n{front_matter}\n---\n\n{text}"


def inline_result_images(result):
    """把 PaddleOCR 返回的图片引用替换为可独立保存的 base64 data URI。"""

    markdown = result["markdown"]
    text = markdown["text"]
    for image_path, image_b64 in (markdown.get("images") or {}).items():
        image_uri = (
            image_b64
            if image_b64.lstrip().startswith("data:image/")
            else f"data:image/png;base64,{image_b64}"
        )
        text = text.replace(f"]({image_path})", f"]({image_uri})")
        text = text.replace(f'src="{image_path}"', f'src="{image_uri}"')
        text = text.replace(f"src='{image_path}'", f"src='{image_uri}'")
    return text


async def paddle_layout(pdf_bytes, client, context):
    """调用 PaddleOCR 布局解析接口，获取逐页文字、媒体和布局结果。"""

    response = await client.post(
        f"{PADDLE_BASE_URL}/layout-parsing",
        json={
            "file": base64.b64encode(pdf_bytes).decode("ascii"),
            "fileType": 0,
            "visualize": True,
            "maxNewTokens": 512,
            "layoutThreshold": 0.7,
            "useDocOrientationClassify": False,
            "useDocUnwarping": False,
            "prettifyMarkdown": False,
        },
    )
    if response.status_code != 200:
        raise RuntimeError(f"{context} layout-parsing 失败：{response.status_code} {response.text}")
    return response.json()["result"]["layoutParsingResults"]


async def paddle_restructure(layout_results, client, concatenate, context):
    """调用 PaddleOCR 页面重构接口，按需合并目录页或保留正文单页。"""

    pages = [
        {
            "prunedResult": result["prunedResult"],
            "markdownImages": result["markdown"].get("images"),
        }
        for result in layout_results
    ]
    response = await client.post(
        f"{PADDLE_BASE_URL}/restructure-pages",
        json={"pages": pages, "concatenatePages": concatenate},
    )
    if response.status_code != 200:
        raise RuntimeError(f"{context} restructure-pages 失败：{response.status_code} {response.text}")
    return response.json()["result"]["layoutParsingResults"]


def save_paddle_layout(layout_results, layout_dir, start_page, pdf_name):
    """按真实 PDF 页码保存 PaddleOCR 的布局 JSON 和可视化图片。"""

    layout_dir.mkdir(parents=True, exist_ok=True)
    for page_number, result in enumerate(layout_results, start=start_page):
        (layout_dir / f"page_{page_number}.json").write_text(
            json.dumps(result["prunedResult"], ensure_ascii=False, indent=2), encoding="utf-8"
        )
        output_images = result.get("outputImages") or {}
        if len(output_images) != 1:
            raise RuntimeError(
                f"{pdf_name} 第 {page_number} 页应返回 1 张 Paddle 布局图，实际 {len(output_images)} 张"
            )
        image_bytes = base64.b64decode(
            paddle._split_data_uri(next(iter(output_images.values()))), validate=True
        )
        suffix = paddle._image_suffix(image_bytes)
        (layout_dir / f"page_{page_number}{suffix}").write_bytes(image_bytes)


def filter_numbered_outline_lines(text):
    """删除目录 OCR 结果中所有不以数字结尾的行，只保留目录条目。"""

    lines = [line for line in text.splitlines() if re.search(r"\d\s*$", line)]
    return "\n".join(lines) + ("\n" if lines else "")


def parse_outline(text, require_headings=True):
    """解析目录标题、书本页码及 Markdown 层级，并拒绝格式异常行。"""

    entries = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        if not line.strip():
            continue
        match = OUTLINE_RE.match(line.strip())
        if not match or (require_headings and not match.group("heading")):
            raise ValueError(f"outline 第 {line_number} 行格式错误：{line}")
        heading = (match.group("heading") or "").strip()
        entries.append(
            {
                "title": match.group("title").strip(),
                "catalog_page": int(match.group("page")),
                "level": len(heading) if heading else None,
            }
        )
    if not entries:
        raise ValueError("outline 中没有识别出目录条目")
    return entries


def validate_outline_response(raw_outline, llm_outline):
    """确认模型只增加目录层级，没有改动标题、页码、数量或顺序。"""

    raw_entries = parse_outline(raw_outline, require_headings=False)
    llm_entries = parse_outline(llm_outline, require_headings=True)
    raw_values = [(normalize_title(e["title"]), e["catalog_page"]) for e in raw_entries]
    llm_values = [(normalize_title(e["title"]), e["catalog_page"]) for e in llm_entries]
    if raw_values != llm_values:
        raise ValueError("LLM 修改了目录标题、页码、数量或顺序")
    return llm_entries


def format_outline(entries):
    """把结构化目录条目序列化为约定的 Markdown 目录格式。"""

    return "\n".join(
        f"{'#' * entry['level']} {entry['title']}……{entry['catalog_page']}" for entry in entries
    ) + "\n"


async def build_outline(raw_outline, model):
    """流式调用目录分层模型，实时打印思考和输出并返回完整目录。"""

    prompt = f"""请根据完整目录判断标题层级，只为每一行添加 Markdown 标题井号。
最高级标题必须使用 ##，下级依次使用 ###、####、#####、######。
不得增加、删除、改写、合并、拆分标题，不得修改页码和顺序。
每行格式必须是：## 标题……页码。不要输出说明、代码围栏或空白段落。

目录：
{raw_outline}"""
    client = AsyncOpenAI(base_url=OUTLINE_BASE_URL, api_key="EMPTY")
    try:
        stream = await client.chat.completions.create(
            model=model,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
            stream=True,
        )
        content_parts = []
        printing_thought = False
        printing_output = False
        async for chunk in stream:
            if not chunk.choices:
                continue
            delta = chunk.choices[0].delta
            extra = getattr(delta, "model_extra", None) or {}
            thought = (
                getattr(delta, "reasoning_content", None)
                or extra.get("reasoning_content")
                or getattr(delta, "reasoning", None)
                or extra.get("reasoning")
            )
            if thought:
                if not printing_thought:
                    print("\n[thought]")
                    printing_thought = True
                print(thought, end="", flush=True)

            content = delta.content
            if content:
                if not printing_output:
                    print("\n\n[output]")
                    printing_output = True
                print(content, end="", flush=True)
                content_parts.append(content)
        print()
        return "".join(content_parts).strip()
    finally:
        await client.close()


def remove_excluded_outline_entries(entries, excluded_tasks, page_offset, book):
    """按正文反选结果删除落入被排除章节页段的目录标题并记录日志。"""

    kept = []
    for entry in entries:
        actual_page = entry["catalog_page"] + page_offset
        excluded = next(
            (
                (task, reason)
                for task, reason in excluded_tasks
                if task["start_page"] <= actual_page < task["end_page"]
            ),
            None,
        )
        if excluded:
            print(
                f"目录反选删除：{book}｜{entry['title']}｜目录页 {entry['catalog_page']}｜"
                f"PDF 页 {actual_page}｜{excluded[1]}"
            )
        else:
            kept.append(entry)
    return kept


async def outline_stage(tasks, infos, model, dry_run):
    """执行完整目录 OCR 和模型分层，产出供人工检查的 outline.md。"""

    all_by_book = group_tasks(tasks)

    if dry_run:
        for book in all_by_book:
            info = infos[book]
            print(f"outline dry-run：{book} 目录 [{info['toc_start']}, {info['toc_end']})")
        return

    async with httpx.AsyncClient(timeout=None) as client:
        for book, book_tasks in all_by_book.items():
            pdf_path = book_tasks[0]["pdf_path"]
            info = infos[book]
            with pdfium.PdfDocument(str(pdf_path)) as pdf:
                if info["toc_end"] > len(pdf) + 1:
                    raise ValueError(f"{book} 的目录范围超出 PDF 页数")

            book_dir = OUTPUT_DIR / book
            final_dir = book_dir / "final_result"
            final_dir.mkdir(parents=True, exist_ok=True)
            pdf_bytes = paddle.extract_pdf_range(pdf_path, info["toc_start"], info["toc_end"])
            layout_results = await paddle_layout(pdf_bytes, client, f"{book} 目录")
            expected = info["toc_end"] - info["toc_start"]
            if len(layout_results) != expected:
                raise RuntimeError(f"{book} 目录应返回 {expected} 页，实际 {len(layout_results)} 页")
            results = await paddle_restructure(layout_results, client, True, f"{book} 目录")
            if len(results) != 1:
                raise RuntimeError(f"{book} 合并目录应返回 1 个结果，实际 {len(results)} 个")
            raw_outline = filter_numbered_outline_lines(inline_result_images(results[0]))
            (book_dir / "outline.md").write_text(raw_outline, encoding="utf-8")

            print(f"开始目录 AI 识别：{book}")
            llm_outline = await build_outline(raw_outline, model)
            try:
                entries = validate_outline_response(raw_outline, llm_outline)
            except ValueError:
                (book_dir / "outline_llm_failed.md").write_text(llm_outline, encoding="utf-8")
                raise
            (final_dir / "outline.md").write_text(format_outline(entries), encoding="utf-8")
            print(f"已生成：{final_dir / 'outline.md'}")
    print("outline 阶段完成，请人工检查 final_result/outline.md 后运行 compare。")


def output_complete(output_dir, layout_dir, page):
    """判断某一页的 Markdown、布局 JSON 和布局图是否已经全部生成。"""

    return (
        (output_dir / f"page_{page}.md").is_file()
        and (layout_dir / f"page_{page}.json").is_file()
        and any(
            (layout_dir / f"page_{page}{suffix}").is_file()
            for suffix in (".jpg", ".png", ".gif", ".webp")
        )
    )


async def convert_paddle_book(book, tasks, client):
    """按表 B 任务用 PaddleOCR 转换一本书的正文页并保存双类输出。"""

    root = OUTPUT_DIR / book / "PaddleOCR"
    output_dir = root / "output"
    layout_dir = root / "layout"
    output_dir.mkdir(parents=True, exist_ok=True)
    layout_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = tasks[0]["pdf_path"]

    for task in tasks:
        pages = range(task["start_page"], task["end_page"])
        if all(output_complete(output_dir, layout_dir, page) for page in pages):
            print(f"跳过已有 PaddleOCR：{book} [{task['start_page']}, {task['end_page']})")
            continue
        pdf_bytes = paddle.extract_pdf_range(pdf_path, task["start_page"], task["end_page"])
        context = f"{book} [{task['start_page']}, {task['end_page']})"
        layout_results = await paddle_layout(pdf_bytes, client, context)
        expected = task["end_page"] - task["start_page"]
        if len(layout_results) != expected:
            raise RuntimeError(f"{context} PaddleOCR 应返回 {expected} 页，实际 {len(layout_results)} 页")
        save_paddle_layout(layout_results, layout_dir, task["start_page"], pdf_path.name)
        results = await paddle_restructure(layout_results, client, False, context)
        if len(results) != expected:
            raise RuntimeError(f"{context} PaddleOCR 重构应返回 {expected} 页，实际 {len(results)} 页")
        for page, result in enumerate(results, start=task["start_page"]):
            text = add_metadata(inline_result_images(result), pdf_path.name, task["metadata"])
            (output_dir / f"page_{page}.md").write_text(text, encoding="utf-8")


def find_zip_file(names, suffix):
    """在 MinerU ZIP 中按后缀定位唯一文件，避免读取错误产物。"""

    matches = [name for name in names if name.lower().endswith(suffix.lower())]
    if len(matches) != 1:
        raise RuntimeError(f"MinerU ZIP 中 {suffix} 应匹配 1 个文件，实际 {len(matches)} 个")
    return matches[0]


def mime_from_bytes(data):
    """根据图片二进制签名确定写入 data URI 所需的 MIME 类型。"""

    return {
        ".jpg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }[paddle._image_suffix(data)]


def inline_mineru_images(markdown, zf):
    """把 MinerU ZIP 内图片嵌入 Markdown，生成无需外部图片文件的结果。"""

    names = zf.namelist()
    for name in names:
        normalized = name.replace("\\", "/")
        if "/images/" not in f"/{normalized}" or normalized.endswith("/"):
            continue
        data = zf.read(name)
        try:
            uri = f"data:{mime_from_bytes(data)};base64,{base64.b64encode(data).decode('ascii')}"
        except ValueError:
            continue
        references = {normalized, normalized.split("/images/", 1)[-1], f"images/{normalized.split('/images/', 1)[-1]}"}
        for reference in references:
            image_pattern = re.compile(rf"!\[([^\]]*)\]\({re.escape(reference)}\)")
            markdown = image_pattern.sub(
                lambda match: (
                    f'<img src="{uri}" alt="{html.escape(match.group(1), quote=True)}">'
                ),
                markdown,
            )
            markdown = markdown.replace(f"]({reference})", f"]({uri})")
            markdown = markdown.replace(f'src="{reference}"', f'src="{uri}"')
            markdown = markdown.replace(f"src='{reference}'", f"src='{uri}'")
    return markdown


async def convert_mineru_page(
    pdf_path, page, metadata, output_dir, layout_dir, zip_dir, client
):
    """调用 MinerU 转换单页，并在校验前保存原始 ZIP 或错误响应。"""

    page_bytes = paddle.extract_pdf_range(pdf_path, page, page + 1)
    response = await client.post(
        MINERU_URL,
        files={"files": (f"page_{page}.pdf", page_bytes, "application/pdf")},
        data={
            "backend": "hybrid-engine",
            "image_analysis": "false",
            "table_enable": "true",
            "formula_enable": "true",
            "return_md": "true",
            "return_middle_json": "false",
            "return_model_output": "true",
            "return_content_list": "false",
            "return_images": "true",
            "response_format_zip": "true",
            "start_page_id": "0",
            "end_page_id": "0",
        },
    )
    if response.status_code != 200:
        error_path = zip_dir / f"page_{page}_error.json"
        error_path.write_bytes(response.content)
        raise RuntimeError(
            f"{pdf_path.name} 第 {page} 页 MinerU 失败：{response.status_code} {response.text}；"
            f"错误响应已保存：{error_path}"
        )

    zip_path = zip_dir / f"page_{page}.zip"
    zip_path.write_bytes(response.content)
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        md_name = find_zip_file(names, ".md")
        markdown = zf.read(md_name).decode("utf-8")
        markdown = inline_mineru_images(markdown, zf)
        json_files = {
            pathlib.PurePosixPath(name).name: json.loads(zf.read(name).decode("utf-8"))
            for name in names
            if name.lower().endswith(".json")
        }

    (layout_dir / f"page_{page}.json").write_text(
        json.dumps(json_files, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / f"page_{page}.md").write_text(
        add_metadata(markdown, pdf_path.name, metadata), encoding="utf-8"
    )


async def convert_mineru_book(book, tasks, client):
    """以五页为一批并发执行 MinerU，批次间串行并汇总失败页面。"""

    root = OUTPUT_DIR / book / "MinerU"
    output_dir = root / "output"
    layout_dir = root / "layout"
    zip_dir = root / "zip"
    output_dir.mkdir(parents=True, exist_ok=True)
    layout_dir.mkdir(parents=True, exist_ok=True)
    zip_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = tasks[0]["pdf_path"]
    pending_pages = []
    for page, metadata in sorted(page_metadata(tasks).items()):
        if (
            (output_dir / f"page_{page}.md").is_file()
            and (layout_dir / f"page_{page}.json").is_file()
            and (zip_dir / f"page_{page}.zip").is_file()
        ):
            print(f"跳过已有 MinerU：{book} 第 {page} 页")
            continue
        pending_pages.append((page, metadata))

    for start in range(0, len(pending_pages), MINERU_BATCH_SIZE):
        batch = pending_pages[start : start + MINERU_BATCH_SIZE]
        page_numbers = "、".join(str(page) for page, _ in batch)
        print(f"MinerU 批次开始：{book}，第 {page_numbers} 页")
        results = await asyncio.gather(
            *(
                convert_mineru_page(
                    pdf_path, page, metadata, output_dir, layout_dir, zip_dir, client
                )
                for page, metadata in batch
            ),
            return_exceptions=True,
        )
        failures = [
            (page, result)
            for (page, _), result in zip(batch, results)
            if isinstance(result, BaseException)
        ]
        if failures:
            details = "；".join(f"第 {page} 页：{error}" for page, error in failures)
            raise RuntimeError(f"{book} MinerU 批次失败：{details}")
        print(f"MinerU 批次完成：{book}，第 {page_numbers} 页")


def contains_media(value):
    """递归检查布局 JSON 是否包含图片、表格或图表类型的内容块。"""

    if isinstance(value, dict):
        for key, item in value.items():
            if key.lower() in {"type", "label", "category", "block_type"}:
                if str(item).lower() in MEDIA_TYPES:
                    return True
            if contains_media(item):
                return True
    elif isinstance(value, list):
        return any(contains_media(item) for item in value)
    return False


def markdown_has_media(text):
    """检查 Markdown/HTML 正文是否存在图片或表格标记。"""

    return bool(re.search(r"<img\b|<table\b|!\[[^\]]*\]\(", text, re.IGNORECASE))


def headings_in_markdown(text):
    """提取页面正文中的 Markdown 标题并规范化，供目录标题核验。"""

    return {
        normalize_title(match.group("title"))
        for line in strip_front_matter(text).splitlines()
        if (match := HEADING_RE.match(line))
    }


def create_review_xlsx(path, rows):
    """创建待人工复核的 Excel，并为处理方式配置固定下拉选项。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人工复核"
    sheet.append(["页数", "原因", "处理方式"])
    for page, reason in rows:
        sheet.append([page, reason, ""])
    validation = DataValidation(
        type="list", formula1=f'"{",".join(REVIEW_OPTIONS)}"', allow_blank=True
    )
    sheet.add_data_validation(validation)
    validation.add(f"C2:C{max(2, len(rows) + 1)}")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{max(1, len(rows) + 1)}"
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 20
    workbook.save(path)


def prepare_compare_outputs(books, overwrite):
    """在 OCR 前检查 compare 目标，并在 overwrite 时清空旧页面和复核表。"""

    targets = []
    for book in books:
        final_dir = OUTPUT_DIR / book / "final_result"
        outline_path = final_dir / "outline.md"
        if not outline_path.is_file():
            raise FileNotFoundError(f"请先完成人工目录检查：{outline_path}")
        review_path = final_dir / "review.xlsx"
        existing_pages = list(final_dir.glob("page_*.md"))
        if (review_path.exists() or existing_pages) and not overwrite:
            raise FileExistsError(f"{book} 已有 compare 结果；如需重建请使用 --overwrite")
        targets.append((review_path, existing_pages))

    if overwrite:
        for review_path, existing_pages in targets:
            review_path.unlink(missing_ok=True)
            for path in existing_pages:
                path.unlink()


def apply_outline_exclusions(excluded, infos):
    """在 compare 开始时从已确认目录中删除落入反选页段的标题。"""

    excluded_by_book = {}
    for task, reason in excluded:
        excluded_by_book.setdefault(task["pdf_path"].stem, []).append((task, reason))

    for book, excluded_tasks in excluded_by_book.items():
        outline_path = OUTPUT_DIR / book / "final_result" / "outline.md"
        if not outline_path.is_file():
            raise FileNotFoundError(f"请先完成人工目录检查：{outline_path}")
        entries = parse_outline(outline_path.read_text(encoding="utf-8"), require_headings=True)
        entries = remove_excluded_outline_entries(
            entries, excluded_tasks, infos[book]["page_offset"], book
        )
        outline_path.write_text(format_outline(entries), encoding="utf-8")


def compare_book(book, tasks, info):
    """按优先级检查双 OCR 页面，将正常页归档并汇总异常页到复核表。"""

    book_dir = OUTPUT_DIR / book
    final_dir = book_dir / "final_result"
    outline_path = final_dir / "outline.md"
    review_path = final_dir / "review.xlsx"

    entries = parse_outline(outline_path.read_text(encoding="utf-8"), require_headings=True)
    titles_by_page = {}
    for entry in entries:
        titles_by_page.setdefault(entry["catalog_page"] + info["page_offset"], []).append(entry)

    paddle_output = book_dir / "PaddleOCR" / "output"
    paddle_layout_dir = book_dir / "PaddleOCR" / "layout"
    mineru_output = book_dir / "MinerU" / "output"
    mineru_layout_dir = book_dir / "MinerU" / "layout"
    issues = []
    for page in sorted(page_metadata(tasks)):
        paddle_md_path = paddle_output / f"page_{page}.md"
        mineru_md_path = mineru_output / f"page_{page}.md"
        paddle_json_path = paddle_layout_dir / f"page_{page}.json"
        mineru_json_path = mineru_layout_dir / f"page_{page}.json"
        for path in (paddle_md_path, mineru_md_path, paddle_json_path, mineru_json_path):
            if not path.is_file():
                raise FileNotFoundError(f"缺少双 OCR 结果：{path}")

        paddle_md = paddle_md_path.read_text(encoding="utf-8")
        mineru_md = mineru_md_path.read_text(encoding="utf-8")
        paddle_layout = json.loads(paddle_json_path.read_text(encoding="utf-8"))
        mineru_layout = json.loads(mineru_json_path.read_text(encoding="utf-8"))

        if (
            markdown_has_media(paddle_md)
            or markdown_has_media(mineru_md)
            or contains_media(paddle_layout)
            or contains_media(mineru_layout)
        ):
            issues.append((page, "含图片表格"))
            continue
        if normalized_markdown(paddle_md) != normalized_markdown(mineru_md):
            issues.append((page, "OCR冲突"))
            continue
        expected_titles = titles_by_page.get(page, [])
        found_titles = headings_in_markdown(paddle_md)
        missing_titles = [
            entry["title"]
            for entry in expected_titles
            if normalize_title(entry["title"]) not in found_titles
        ]
        if missing_titles:
            print(f"标题查找失败：{book} 第 {page} 页：{', '.join(missing_titles)}")
            issues.append((page, "标题查找失败"))
            continue
        if not normalized_markdown(paddle_md).strip():
            issues.append((page, "空白页面"))
            continue
        shutil.copy2(mineru_md_path, final_dir / f"page_{page}.md")

    create_review_xlsx(review_path, issues)
    print(f"已生成：{review_path}（{len(issues)} 个待复核页面）")


async def compare_stage(included, excluded, infos, overwrite, dry_run):
    """执行入选正文页的双 OCR 转换、自动对比和人工复核表生成。"""

    included_tasks = [task for task, _ in included]
    by_book = group_tasks(included_tasks)
    if dry_run:
        for book, book_tasks in by_book.items():
            print(f"compare dry-run：{book}，{len(page_metadata(book_tasks))} 页")
        return

    apply_outline_exclusions(excluded, infos)
    prepare_compare_outputs(by_book, overwrite)
    async with httpx.AsyncClient(timeout=None) as client:
        for book, book_tasks in by_book.items():
            await convert_paddle_book(book, book_tasks, client)
            await convert_mineru_book(book, book_tasks, client)
            compare_book(book, book_tasks, infos[book])
    print("compare 阶段完成，请填写 final_result/review.xlsx 后运行 finalize。")


def read_review(path):
    """读取人工复核表，并校验每个异常页都填写了合法处理方式。"""

    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [sheet.cell(1, column).value for column in range(1, 4)]
    if headers != ["页数", "原因", "处理方式"]:
        raise ValueError(f"复核表表头错误：{headers}")
    rows = []
    for page, reason, decision in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if page is None:
            continue
        decision = (decision or "").strip()
        if decision and decision not in REVIEW_OPTIONS:
            raise ValueError(f"第 {page} 页处理方式无效：{decision}")
        rows.append((int(page), str(reason), decision))
    return rows


def replace_heading(text, title, level):
    """在单页 Markdown 中定位指定标题并替换成目录确认后的层级。"""

    lines = text.splitlines()
    target = normalize_title(title)
    for index, line in enumerate(lines):
        match = HEADING_RE.match(line)
        if match and normalize_title(match.group("title")) == target:
            lines[index] = f"{match.group('indent')}{'#' * level} {match.group('title')}"
            return "\n".join(lines) + ("\n" if text.endswith("\n") else ""), True
    return text, False


def apply_outline_levels(final_dir, outline_entries, page_offset, deleted_pages):
    """把最终目录层级批量应用到已确认保留的正文页面。"""

    pending = {}
    missing = []
    for entry in outline_entries:
        page = entry["catalog_page"] + page_offset
        if page in deleted_pages:
            continue
        path = final_dir / f"page_{page}.md"
        if not path.is_file():
            continue
        text = pending.get(path, path.read_text(encoding="utf-8"))
        text, found = replace_heading(text, entry["title"], entry["level"])
        if not found:
            missing.append(f"第 {page} 页：{entry['title']}")
        pending[path] = text
    if missing:
        raise ValueError("最终文件标题查找失败：\n" + "\n".join(missing))
    for path, text in pending.items():
        path.write_text(text, encoding="utf-8")


def finalize_book(book, tasks, info):
    """应用人工复核决策、修正标题层级并统计一本书的双 OCR 失误率。"""

    book_dir = OUTPUT_DIR / book
    final_dir = book_dir / "final_result"
    review_path = final_dir / "review.xlsx"
    outline_path = final_dir / "outline.md"
    if not review_path.is_file() or not outline_path.is_file():
        raise FileNotFoundError(f"{book} 缺少 review.xlsx 或 outline.md")

    review_rows = read_review(review_path)
    paddle_errors = 0
    mineru_errors = 0
    deleted_pages = set()
    skipped_pages = set()
    for page, reason, decision in review_rows:
        if not decision:
            skipped_pages.add(page)
            continue
        destination = final_dir / f"page_{page}.md"
        paddle_page = book_dir / "PaddleOCR" / "output" / f"page_{page}.md"
        mineru_page = book_dir / "MinerU" / "output" / f"page_{page}.md"
        if decision == "保留PaddleOCR":
            shutil.copy2(paddle_page, destination)
            mineru_errors += 1
        elif decision == "保留MinerU":
            shutil.copy2(mineru_page, destination)
            paddle_errors += 1
        elif decision == "手动merge":
            if not destination.is_file():
                raise FileNotFoundError(f"第 {page} 页选择手动merge，但最终文件不存在：{destination}")
            paddle_errors += 1
            mineru_errors += 1
        elif decision == "删除该页":
            destination.unlink(missing_ok=True)
            deleted_pages.add(page)
            if reason != "空白页面":
                paddle_errors += 1
                mineru_errors += 1
        elif decision == "两种都可":
            shutil.copy2(mineru_page, destination)

    if skipped_pages:
        pages = "、".join(str(page) for page in sorted(skipped_pages))
        print(f"处理方式为空，已跳过：{tasks[0]['pdf_path'].name} 第 {pages} 页")

    processed_pages = set(page_metadata(tasks))
    missing_final = sorted(
        page
        for page in processed_pages - deleted_pages - skipped_pages
        if not (final_dir / f"page_{page}.md").is_file()
    )
    if missing_final:
        raise FileNotFoundError(f"最终结果缺少页面：{missing_final}")

    outline_entries = parse_outline(outline_path.read_text(encoding="utf-8"), require_headings=True)
    apply_outline_levels(final_dir, outline_entries, info["page_offset"], deleted_pages)

    denominator = len(processed_pages - skipped_pages)
    result = {
        "书名": book,
        "实际处理页数": denominator,
        "目录页数": info["toc_end"] - info["toc_start"],
        "删除页数": len(deleted_pages),
        "PaddleOCR": {
            "错误页数": paddle_errors,
            "错误率": round(paddle_errors / denominator, 6) if denominator else 0,
        },
        "MinerU": {
            "错误页数": mineru_errors,
            "错误率": round(mineru_errors / denominator, 6) if denominator else 0,
        },
    }
    (final_dir / "result.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"已完成：{final_dir}")


def finalize_stage(included, infos, dry_run):
    """按书执行人工复核结果落盘和最终统计。"""

    by_book = group_tasks([task for task, _ in included])
    if dry_run:
        for book, book_tasks in by_book.items():
            print(f"finalize dry-run：{book}，{len(page_metadata(book_tasks))} 页")
        return
    for book, book_tasks in by_book.items():
        finalize_book(book, book_tasks, infos[book])


def resolve_config(args):
    """用 dual 自身的全局默认值解析反选和 dry-run 参数。"""

    pair_values = EXCLUDE_PAIRS if args.exclude_pair is None else args.exclude_pair
    return {
        "structures": set(
            EXCLUDE_STRUCTURES if args.exclude_structure is None else args.exclude_structure
        ),
        "scenes": set(EXCLUDE_SCENES if args.exclude_scene is None else args.exclude_scene),
        "branches": set(
            EXCLUDE_BRANCHES if args.exclude_branch is None else args.exclude_branch
        ),
        "pairs": paddle.parse_pairs(pair_values),
        "dry_run": DRY_RUN if args.dry_run is None else args.dry_run,
    }


def build_parser():
    """定义三阶段流程及反选、覆盖和试运行命令行参数。"""

    parser = argparse.ArgumentParser(description="目录分层、双 OCR 对比与人工复核流程")
    parser.add_argument("stage", choices=("outline", "compare", "finalize"))
    parser.add_argument("--book", nargs="*", default=None, help="只处理指定重命名书名")
    parser.add_argument("--outline-model", default=OUTLINE_MODEL)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--exclude-structure", nargs="*", default=None)
    parser.add_argument("--exclude-scene", nargs="*", default=None)
    parser.add_argument("--exclude-branch", nargs="*", default=None)
    parser.add_argument("--exclude-pair", nargs="*", default=None, metavar="场景-子分支")
    parser.add_argument(
        "--dry-run", action=argparse.BooleanOptionalAction, default=None, help="只校验，不调用接口"
    )
    return parser


async def main(args):
    """加载配置与表格任务，并调度 outline、compare 或 finalize 阶段。"""

    config = resolve_config(args)
    tasks, missing_books = load_b_tasks(set(args.book) if args.book else None)
    if not tasks:
        print(f"没有可处理的 PDF；已跳过 {len(missing_books)} 本")
        return

    unknown_filters = paddle.find_unknown_filters(tasks, config)
    if unknown_filters:
        if config["dry_run"]:
            paddle.print_unknown_filters(unknown_filters)
        elif not paddle.confirm_unknown_filters(unknown_filters):
            print("已取消")
            return
    included, excluded = paddle.filter_tasks(tasks, config)
    paddle.print_summary(included, excluded, config)
    book_names = {task["pdf_path"].stem for task in tasks}
    infos = load_book_infos(book_names)

    if args.stage == "outline":
        await outline_stage(tasks, infos, args.outline_model, config["dry_run"])
    elif args.stage == "compare":
        await compare_stage(included, excluded, infos, args.overwrite, config["dry_run"])
    else:
        finalize_stage(included, infos, config["dry_run"])


if __name__ == "__main__":
    asyncio.run(main(build_parser().parse_args()))
